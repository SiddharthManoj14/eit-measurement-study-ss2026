import os
import numpy as np
import pandas as pd

from openpyxl.styles import PatternFill


def data_txt_to_excel(folder_path, excel_file, high_value_limit=9):
    """
    Reads all .txt files in folder_path and exports them to Excel.

    Excel format:
    rows    = voltage indices
    columns = Measurement_1, Measurement_2, ...

    Cells with values greater than high_value_limit
    are highlighted in yellow.
    """

    median_voltages = {}

    yellow_fill = PatternFill(
        start_color="FFF2CC",
        end_color="FFF2CC",
        fill_type="solid"
    )

    output_folder = os.path.dirname(excel_file)

    if output_folder and not os.path.exists(output_folder):
        os.makedirs(output_folder)

    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        files = os.listdir(folder_path)

        for file in files:
            if not file.endswith(".txt"):
                continue

            file_path = os.path.join(folder_path, file)
            print("Reading:", file_path)

            values_all = []

            with open(file_path, "r") as f:
                lines = f.readlines()

            measurement_number = 1

            for line in lines:
                if not line.startswith("I"):
                    continue

                str_data = line.split()
                float_data = []

                for value in str_data:
                    if value == "I" or value == "got:":
                        continue

                    try:
                        float_data.append(float(value))
                    except ValueError:
                        print(
                            "Could not convert:",
                            value,
                            "in measurement",
                            measurement_number
                        )

                count = len(float_data)

                if count == 208:
                    values_all.append(float_data)

                elif count > 208:
                    print(
                        "More than 208 values found at measurement:",
                        measurement_number,
                        "count:",
                        count
                    )

                else:
                    print(
                        "Less than 208 values found at measurement:",
                        measurement_number,
                        "count:",
                        count
                    )

                measurement_number += 1

            data = np.array(values_all)

            if data.size == 0:
                print("No valid data found in:", file)
                continue

            median_values = np.median(data, axis=0)
            median_voltages[file_path] = median_values

            # Only mark very high values as outliers.
            # Example: value > 9
            outlier_mask = data > high_value_limit

            # Keep original Excel format:
            # data shape:
            # rows    = measurements
            # columns = 208 voltage values
            #
            # Excel shape after transpose:
            # rows    = voltage indices
            # columns = Measurement_1, Measurement_2, ...
            df = pd.DataFrame(data).T

            df.columns = [
                "Measurement_" + str(i + 1)
                for i in range(df.shape[1])
            ]

            df["Median"] = median_values

            sheet_name = file[:-4][:31]

            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

            worksheet = writer.book[sheet_name]

            # Mapping:
            #
            # data[measurement_index, voltage_index]
            #
            # becomes Excel:
            #
            # row = voltage_index + 2
            # col = measurement_index + 1
            #
            # +2 because Excel row 1 is the header.
            # +1 because Excel columns start from 1.
            for measurement_idx in range(outlier_mask.shape[0]):
                for voltage_idx in range(outlier_mask.shape[1]):
                    if outlier_mask[measurement_idx, voltage_idx]:
                        excel_row = voltage_idx + 2
                        excel_col = measurement_idx + 1

                        worksheet.cell(
                            row=excel_row,
                            column=excel_col
                        ).fill = yellow_fill

    return median_voltages


folder = (
    r"C:\Users\manoj\OneDrive\Desktop"
    r"\measurements_29042026\Adjacent"
    r"\rectangle\non-conductive"
)

excel_output = (
    r"C:\Users\manoj\OneDrive\Desktop"
    r"\measurements_29042026\collected data test"
    r"\rectangular_tank_circular_object_highlighted.xlsx"
)

mv = data_txt_to_excel(
    folder_path=folder,
    excel_file=excel_output,
    high_value_limit=9
)

print("Excel file created at:")
print(excel_output)